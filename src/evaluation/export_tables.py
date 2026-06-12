"""
src/evaluation/export_tables.py
================================

Thesis-ready table export in CSV, Excel (xlsx), and LaTeX (booktabs) formats.

Generates all tables that appear in the dissertation from the data structures
produced by the rest of the pipeline:

Tables produced
---------------
``pareto_table``
    Pareto-optimal candidates with objectives and key architecture /
    quantisation parameters.  Rows sorted by AUROC descending.

``per_category_table``
    Per-MVTec-category AUROC (and optionally PR-AUC, pixel AUROC) for one or
    more model configurations.  Includes mean ± std footer rows.

``statistical_table``
    Pairwise Wilcoxon results (W, p, r) and Friedman χ² summary, produced
    from a :class:`src.evaluation.statistics.ModelComparisonReport`.

``architecture_table``
    Architecture + quantisation hyperparameters for selected candidates:
    family, depth, width, kernel, skip, W_bits, A_bits, params (M), MACs (G).

``tracking_table``
    Per-scenario tracking performance: IoU, precision, recall, F1, failure
    rate, FPS.

``repeatability_table``
    Across-seed reproducibility: mean, std, ICC, CV, SEM, MDC95, per metric.

``to_latex``
    Generic converter — accepts a list of row dicts or 2-D list plus headers
    and returns a complete LaTeX ``table`` environment string.

``export_all_tables``
    Orchestrator — call once from ``main_report.py`` with all data in hand.

Output layout::

    <output_dir>/
    ├── csv/
    │   ├── pareto.csv
    │   ├── per_category.csv
    │   └── …
    ├── latex/
    │   ├── pareto.tex
    │   ├── per_category.tex
    │   └── …
    └── excel/
        └── all_tables.xlsx   (one sheet per table)

Assumptions
-----------
- Candidate dicts carry at minimum: ``auroc``, ``latency_ms``,
  ``peak_ram_mb``, ``energy_mj``, and optionally ``is_pareto``,
  ``arch_family``, ``depth``, ``width``, ``weight_bits``, ``act_bits``,
  ``symmetric``, ``per_channel``, ``n_params_m``, ``macs_g``.
- ``openpyxl`` is preferred for Excel; the module falls back to CSV-only
  output when it is absent.
- ``pandas`` is used for CSV/Excel construction when available; a pure-Python
  fallback handles the CSV path otherwise.
"""

from __future__ import annotations

import csv
import io
import logging
import math
import re
import warnings
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Sequence

LOG = logging.getLogger(__name__)

try:
    import pandas as _pd  # type: ignore
    _HAVE_PANDAS = True
except ImportError:  # pragma: no cover
    _pd = None  # type: ignore
    _HAVE_PANDAS = False

try:
    import openpyxl  # type: ignore
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    _HAVE_OPENPYXL = True
except ImportError:  # pragma: no cover
    openpyxl = None  # type: ignore
    _HAVE_OPENPYXL = False

__all__ = [
    "TableConfig",
    "to_latex",
    "export_pareto_table",
    "export_per_category_table",
    "export_statistical_table",
    "export_architecture_table",
    "export_tracking_table",
    "export_repeatability_table",
    "export_all_tables",
]


# ---------------------------------------------------------------------------
# Column metadata registry
# ---------------------------------------------------------------------------

#: Centralised display properties for every known column key.
COLUMN_DEFS: dict[str, dict[str, Any]] = {
    # Objectives
    "auroc":        {"header": "AUROC",         "unit": "",         "better": "higher", "fmt": ".4f", "align": "r"},
    "image_auroc":  {"header": "AUROC",         "unit": "",         "better": "higher", "fmt": ".4f", "align": "r"},
    "pixel_auroc":  {"header": "Pixel AUROC",   "unit": "",         "better": "higher", "fmt": ".4f", "align": "r"},
    "pr_auc":       {"header": "PR-AUC",        "unit": "",         "better": "higher", "fmt": ".4f", "align": "r"},
    "pixel_pr_auc": {"header": "Pixel PR-AUC",  "unit": "",         "better": "higher", "fmt": ".4f", "align": "r"},
    "latency_ms":   {"header": "Latency",       "unit": "ms",       "better": "lower",  "fmt": ".2f", "align": "r"},
    "peak_ram_mb":  {"header": "Peak RAM",      "unit": "MB",       "better": "lower",  "fmt": ".1f", "align": "r"},
    "energy_mj":    {"header": "Energy",        "unit": "mJ",       "better": "lower",  "fmt": ".3f", "align": "r"},
    # Classification
    "f1":           {"header": "F1",            "unit": "",         "better": "higher", "fmt": ".4f", "align": "r"},
    "accuracy":     {"header": "Accuracy",      "unit": "",         "better": "higher", "fmt": ".4f", "align": "r"},
    "precision":    {"header": "Precision",     "unit": "",         "better": "higher", "fmt": ".4f", "align": "r"},
    "recall":       {"header": "Recall",        "unit": "",         "better": "higher", "fmt": ".4f", "align": "r"},
    # Architecture
    "arch_family":  {"header": "Family",        "unit": "",         "better": None,     "fmt": "s",   "align": "l"},
    "depth":        {"header": "Depth",         "unit": "",         "better": None,     "fmt": "d",   "align": "r"},
    "width":        {"header": "Width",         "unit": "",         "better": None,     "fmt": "d",   "align": "r"},
    "weight_bits":  {"header": r"$b_w$",        "unit": "bit",      "better": None,     "fmt": "d",   "align": "r"},
    "act_bits":     {"header": r"$b_a$",        "unit": "bit",      "better": None,     "fmt": "d",   "align": "r"},
    "symmetric":    {"header": "Sym.",          "unit": "",         "better": None,     "fmt": "s",   "align": "c"},
    "per_channel":  {"header": "Per-ch.",       "unit": "",         "better": None,     "fmt": "s",   "align": "c"},
    "n_params_m":   {"header": "Params",        "unit": "M",        "better": "lower",  "fmt": ".2f", "align": "r"},
    "macs_g":       {"header": "MACs",          "unit": "G",        "better": "lower",  "fmt": ".2f", "align": "r"},
    # Statistical
    "statistic":    {"header": "Statistic",     "unit": "",         "better": None,     "fmt": ".4f", "align": "r"},
    "p_value":      {"header": r"$p$-value",    "unit": "",         "better": None,     "fmt": ".4f", "align": "r"},
    "effect_size":  {"header": r"$r$",          "unit": "",         "better": None,     "fmt": ".3f", "align": "r"},
    "significant":  {"header": "Sig.",          "unit": "",         "better": None,     "fmt": "s",   "align": "c"},
    # Repeatability
    "icc":          {"header": "ICC(2,1)",      "unit": "",         "better": "higher", "fmt": ".4f", "align": "r"},
    "cv":           {"header": "CV",            "unit": r"\%",      "better": "lower",  "fmt": ".2f", "align": "r"},
    "sem":          {"header": "SEM",           "unit": "",         "better": "lower",  "fmt": ".4f", "align": "r"},
    "mdc95":        {"header": "MDC$_{95}$",    "unit": "",         "better": "lower",  "fmt": ".4f", "align": "r"},
    # Tracking
    "iou":          {"header": "IoU",           "unit": "",         "better": "higher", "fmt": ".4f", "align": "r"},
    "failure_rate": {"header": "Fail rate",     "unit": r"\%",      "better": "lower",  "fmt": ".2f", "align": "r"},
    "fps":          {"header": "FPS",           "unit": "",         "better": "higher", "fmt": ".1f", "align": "r"},
}


# ---------------------------------------------------------------------------
# Configuration
# ---------------------------------------------------------------------------

@dataclass
class TableConfig:
    """Visual and formatting options shared across all table exports."""
    # Number formats (fallback when key not in COLUMN_DEFS)
    fmt_float:           str  = ".4f"
    fmt_int:             str  = "d"
    missing:             str  = r"{\textemdash}"  # LaTeX em-dash for None
    missing_csv:         str  = ""

    # LaTeX options
    bold_best:           bool = True    # bold the best cell in each column
    table_pos:           str  = "htbp"
    font_size_cmd:       str  = ""      # "" | r"\small" | r"\footnotesize"
    booktabs:            bool = True    # use \toprule / \midrule / \bottomrule
    alternating_rows:    bool = False   # \rowcolor{gray!10} on odd rows
    siunitx:             bool = False   # use S column type for number alignment
    add_units_row:       bool = True    # print units in a sub-header row
    caption_prefix:      str  = "Table: "
    add_label:           bool = True

    # Excel options
    freeze_header:       bool = True
    header_fill_hex:     str  = "1F4E79"  # dark blue header
    best_fill_hex:       str  = "D9EAD3"  # light green for best values
    alternating_fill_hex: str = "F2F2F2"

    # Output formats
    export_csv:          bool = True
    export_latex:        bool = True
    export_excel:        bool = True


# ---------------------------------------------------------------------------
# LaTeX helpers
# ---------------------------------------------------------------------------

_LATEX_SPECIAL = str.maketrans({
    "&":  r"\&",
    "%":  r"\%",
    "$":  r"\$",
    "#":  r"\#",
    "_":  r"\_",
    "{":  r"\{",
    "}":  r"\}",
    "~":  r"\textasciitilde{}",
    "^":  r"\textasciicircum{}",
    "\\": r"\textbackslash{}",
})


def _latex_escape(s: str) -> str:
    """Escape a plain string for safe use in LaTeX (does not escape math mode)."""
    # Leave strings that already contain LaTeX commands untouched.
    if re.search(r"\\[a-zA-Z{]", s) or "$" in s:
        return s
    return s.translate(_LATEX_SPECIAL)


def _bold_latex(s: str) -> str:
    return rf"\textbf{{{s}}}"


def _fmt_num(val: Any, fmt: str, missing: str = "—") -> str:
    """Format a scalar value as a string."""
    if val is None or (isinstance(val, float) and math.isnan(val)):
        return missing
    if fmt == "s":
        return str(val)
    if fmt == "d":
        try:
            return str(int(round(float(val))))
        except (ValueError, TypeError):
            return str(val)
    try:
        return format(float(val), fmt)
    except (ValueError, TypeError):
        return str(val)


def _col_header(key: str) -> str:
    defn = COLUMN_DEFS.get(key, {})
    return str(defn.get("header", key.replace("_", " ").title()))


def _col_unit(key: str) -> str:
    defn = COLUMN_DEFS.get(key, {})
    u = str(defn.get("unit", ""))
    return f"[{u}]" if u else ""


def _col_align(key: str) -> str:
    return str(COLUMN_DEFS.get(key, {}).get("align", "r"))


def _col_fmt(key: str, cfg: TableConfig) -> str:
    return str(COLUMN_DEFS.get(key, {}).get("fmt", cfg.fmt_float))


def _col_better(key: str) -> str | None:
    return COLUMN_DEFS.get(key, {}).get("better")


def _best_row_indices(rows: list[list[Any]],
                      col_idx: int,
                      better: str | None) -> set[int]:
    """Return indices of rows that carry the best value in column *col_idx*."""
    if better is None:
        return set()
    vals: list[tuple[int, float]] = []
    for i, row in enumerate(rows):
        v = row[col_idx]
        if v is not None and isinstance(v, (int, float)) and not math.isnan(float(v)):
            vals.append((i, float(v)))
    if not vals:
        return set()
    best_val = max(v for _, v in vals) if better == "higher" else min(v for _, v in vals)
    return {i for i, v in vals if math.isclose(v, best_val, rel_tol=1e-9)}


# ---------------------------------------------------------------------------
# Generic LaTeX table builder
# ---------------------------------------------------------------------------

def to_latex(
        data: list[dict[str, Any]] | list[list[Any]],
        headers: list[str] | None = None,
        *,
        col_keys: list[str] | None = None,
        caption: str = "",
        label: str = "",
        cfg: TableConfig | None = None,
        footer_rows: list[list[Any]] | None = None,
) -> str:
    """Convert tabular data to a complete LaTeX ``table`` environment.

    Parameters
    ----------
    data:
        Either a list of row dicts (keys = column identifiers) or a 2-D list
        of raw values.  When dicts are supplied, ``col_keys`` selects and
        orders the columns.
    headers:
        Display column headers.  When ``data`` is a list-of-dicts and
        ``col_keys`` is given, headers default to the ``COLUMN_DEFS``
        display names.
    col_keys:
        Ordered column identifiers (for COLUMN_DEFS lookup).  Required when
        ``data`` is a list of dicts; ignored for 2-D list input.
    caption, label:
        LaTeX ``\caption{}`` and ``\label{tab:...}``.
    cfg:
        Formatting configuration.
    footer_rows:
        Extra rows appended below a midrule (e.g., mean/std summary).

    Returns
    -------
    str
        Complete LaTeX ``table`` environment.
    """
    cfg = cfg or TableConfig()

    # Normalise to list-of-lists
    if data and isinstance(data[0], dict):
        if col_keys is None:
            col_keys = list(data[0].keys())
        rows: list[list[Any]] = [[row.get(k) for k in col_keys] for row in data]
        if headers is None:
            headers = [_col_header(k) for k in col_keys]
    else:
        rows = [list(r) for r in data]  # type: ignore[arg-type]
        col_keys = col_keys or [str(i) for i in range(len(rows[0]) if rows else 0)]
        if headers is None:
            headers = col_keys

    n_cols = len(headers)

    # Column format string
    if cfg.siunitx:
        col_fmt_str = " ".join(
            "S" if _col_better(k) is not None or
                   COLUMN_DEFS.get(k, {}).get("fmt", "s") not in ("s", "d")
            else _col_align(k)
            for k in (col_keys or ["l"] * n_cols)
        )
    else:
        col_fmt_str = " ".join(_col_align(k) for k in (col_keys or ["l"] * n_cols))

    # Identify best cells per column
    best_by_col: list[set[int]] = []
    for ci, key in enumerate(col_keys or []):
        best_by_col.append(
            _best_row_indices(rows, ci, _col_better(key)) if cfg.bold_best else set()
        )

    def _render_row(row: list[Any], row_idx: int | None, is_footer: bool = False) -> str:
        cells: list[str] = []
        for ci, (val, key) in enumerate(zip(row, col_keys or [""] * n_cols)):
            fmt = _col_fmt(key, cfg)
            cell_str = _fmt_num(val, fmt, cfg.missing)
            # Bold best (only in data rows, not footer)
            if (not is_footer and cfg.bold_best
                    and row_idx is not None
                    and row_idx in best_by_col[ci]):
                cell_str = _bold_latex(cell_str)
            # Escape plain strings
            if not (cell_str.startswith("\\") or "$" in cell_str):
                cell_str = _latex_escape(cell_str)
            cells.append(cell_str)
        sep = " & "
        return sep.join(cells) + r" \\"

    lines: list[str] = []

    if cfg.font_size_cmd:
        lines.append(cfg.font_size_cmd)

    # tabular
    top    = r"\toprule"    if cfg.booktabs else r"\hline"
    mid    = r"\midrule"    if cfg.booktabs else r"\hline"
    bottom = r"\bottomrule" if cfg.booktabs else r"\hline"

    lines.append(rf"\begin{{tabular}}{{{col_fmt_str}}}")
    lines.append(top)

    # Header row(s)
    header_cells = [_latex_escape(_col_header(k)) for k in (col_keys or headers)]
    lines.append(" & ".join(header_cells) + r" \\")

    if cfg.add_units_row and col_keys:
        unit_cells = [_latex_escape(_col_unit(k)) for k in col_keys]
        if any(u for u in unit_cells):
            lines.append(" & ".join(unit_cells) + r" \\")

    lines.append(mid)

    # Data rows
    for ri, row in enumerate(rows):
        prefix = r"\rowcolor{gray!10}" if cfg.alternating_rows and ri % 2 == 1 else ""
        rendered = _render_row(row, ri)
        lines.append(prefix + rendered)

    # Footer rows (summary stats, mean/std)
    if footer_rows:
        lines.append(mid)
        for row in footer_rows:
            lines.append(_render_row(row, None, is_footer=True))

    lines.append(bottom)
    lines.append(r"\end{tabular}")

    # Wrap in table environment
    wrapped: list[str] = [rf"\begin{{table}}[{cfg.table_pos}]", r"\centering"]
    if cfg.font_size_cmd:
        wrapped.append(cfg.font_size_cmd)
    if caption:
        wrapped.append(rf"\caption{{{_latex_escape(caption)}}}")
    if label and cfg.add_label:
        clean_label = re.sub(r"[^a-zA-Z0-9:_-]", "_", label)
        wrapped.append(rf"\label{{tab:{clean_label}}}")
    wrapped.extend(lines)
    wrapped.append(r"\end{table}")

    return "\n".join(wrapped)


# ---------------------------------------------------------------------------
# CSV / Excel writers
# ---------------------------------------------------------------------------

def _write_csv(rows: list[list[Any]],
               headers: list[str],
               path: Path,
               missing: str = "") -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.writer(fh)
        writer.writerow(headers)
        for row in rows:
            writer.writerow([
                "" if (v is None or (isinstance(v, float) and math.isnan(v)))
                else v
                for v in row
            ])
    return path


def _write_excel_sheet(wb,  # openpyxl.Workbook
                       sheet_name: str,
                       rows: list[list[Any]],
                       headers: list[str],
                       col_keys: list[str] | None,
                       cfg: TableConfig,
                       best_by_col: list[set[int]] | None = None) -> None:
    """Write one sheet into an open openpyxl workbook."""
    ws = wb.create_sheet(title=sheet_name[:31])  # Excel 31-char limit

    hdr_font   = Font(color="FFFFFF", bold=True)
    hdr_fill   = PatternFill("solid", fgColor=cfg.header_fill_hex)
    best_fill  = PatternFill("solid", fgColor=cfg.best_fill_hex)
    alt_fill   = PatternFill("solid", fgColor=cfg.alternating_fill_hex)
    thin       = Side(style="thin")
    border     = Border(bottom=thin)
    center_aln = Alignment(horizontal="center", vertical="center")

    # Header row
    for ci, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=ci, value=hdr)
        cell.font = hdr_fill and hdr_font
        cell.fill = hdr_fill
        cell.alignment = center_aln

    if cfg.freeze_header:
        ws.freeze_panes = "A2"

    # Data rows
    for ri, row in enumerate(rows, start=2):
        for ci, val in enumerate(row, start=1):
            v = None if (isinstance(val, float) and math.isnan(val)) else val
            cell = ws.cell(row=ri, column=ci, value=v)
            if cfg.alternating_rows and (ri % 2 == 0):
                cell.fill = alt_fill
            if best_by_col and (ri - 2) in best_by_col[ci - 1]:
                cell.fill = best_fill
                cell.font = Font(bold=True)

    # Auto-fit column widths (heuristic)
    for ci, hdr in enumerate(headers, start=1):
        max_len = len(str(hdr))
        for row in rows:
            v = row[ci - 1]
            max_len = max(max_len, len(str(v) if v is not None else ""))
        ws.column_dimensions[get_column_letter(ci)].width = min(max_len + 4, 40)


# ---------------------------------------------------------------------------
# Individual table generators
# ---------------------------------------------------------------------------

def export_pareto_table(
        candidates: list[dict[str, Any]],
        *,
        pareto_only: bool = True,
        col_keys: list[str] | None = None,
        output_dir: Path | str = Path("."),
        caption: str = "Pareto-optimal model configurations.",
        label: str = "pareto_front",
        cfg: TableConfig | None = None,
) -> dict[str, Path]:
    """Export the Pareto-front candidate table.

    Parameters
    ----------
    candidates:
        Full population list; filtered to ``is_pareto == True`` when
        ``pareto_only=True``.
    col_keys:
        Columns to include.  Defaults to objectives + key arch/quant params.
    output_dir:
        Root output directory.
    caption, label:
        LaTeX caption and label suffix.
    cfg:
        Formatting configuration.

    Returns
    -------
    dict mapping format name → file Path.
    """
    cfg = cfg or TableConfig()
    out = Path(output_dir)

    subset = (
        [c for c in candidates if c.get("is_pareto")]
        if pareto_only else candidates
    )
    # Sort by AUROC descending
    subset = sorted(subset, key=lambda c: float(c.get("auroc") or 0), reverse=True)

    default_keys = [
        "arch_family", "depth", "width",
        "weight_bits", "act_bits",
        "auroc", "latency_ms", "peak_ram_mb", "energy_mj",
        "n_params_m", "macs_g",
    ]
    keys = col_keys or [k for k in default_keys if any(k in c for c in subset)]
    headers = [_col_header(k) for k in keys]
    rows = [[c.get(k) for k in keys] for c in subset]

    best_by_col = [_best_row_indices(rows, ci, _col_better(k))
                   for ci, k in enumerate(keys)]

    saved: dict[str, Path] = {}

    if cfg.export_csv:
        p = _write_csv(rows, headers, out / "csv" / "pareto.csv", cfg.missing_csv)
        saved["csv"] = p

    if cfg.export_latex:
        tex = to_latex(subset, col_keys=keys,
                       caption=caption, label=label, cfg=cfg)
        p = out / "latex" / "pareto.tex"
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(tex, encoding="utf-8")
        saved["latex"] = p
        LOG.info("LaTeX table → %s", p)

    saved["_rows"]        = rows         # type: ignore[assignment]
    saved["_headers"]     = headers      # type: ignore[assignment]
    saved["_keys"]        = keys         # type: ignore[assignment]
    saved["_best_by_col"] = best_by_col  # type: ignore[assignment]
    return saved


def export_per_category_table(
        results_by_model: dict[str, dict[str, dict[str, Any]]],
        categories: list[str] | None = None,
        *,
        metric: str = "auroc",
        output_dir: Path | str = Path("."),
        caption: str | None = None,
        label: str = "per_category",
        cfg: TableConfig | None = None,
) -> dict[str, Path]:
    """Export a per-category metric table with one column per model.

    Parameters
    ----------
    results_by_model:
        ``{model_name: {category: metrics_dict}}``.  Each ``metrics_dict``
        must carry at least the requested ``metric`` key.
    categories:
        Row order.  Defaults to sorted union of all categories found.
    metric:
        Key to extract (``"auroc"``, ``"pixel_auroc"``, ``"pr_auc"`` …).
    output_dir, caption, label, cfg:
        Standard export parameters.
    """
    cfg = cfg or TableConfig()
    out = Path(output_dir)

    model_names = list(results_by_model.keys())
    if not categories:
        cats: set[str] = set()
        for m in results_by_model.values():
            cats |= m.keys()
        categories = sorted(cats)

    import math as _math

    rows: list[list[Any]] = []
    for cat in categories:
        row: list[Any] = [cat]
        for mn in model_names:
            v = results_by_model[mn].get(cat, {}).get(metric)
            row.append(v)
        rows.append(row)

    # Summary footer: mean ± std
    for stat_name, fn in [("Mean", lambda vs: sum(vs) / len(vs) if vs else None),
                           ("Std",  lambda vs: (
                               math.sqrt(sum((v - sum(vs)/len(vs))**2 for v in vs) / max(len(vs)-1,1))
                               if len(vs) > 1 else 0.0))]:
        footer: list[Any] = [stat_name]
        for col_i in range(1, len(model_names) + 1):
            col_vals = [r[col_i] for r in rows
                        if r[col_i] is not None and not _math.isnan(float(r[col_i] or "nan"))]
            footer.append(fn(col_vals))  # type: ignore[arg-type]
        rows.append(footer)  # will be split out below

    data_rows   = rows[:-2]
    footer_rows = rows[-2:]

    headers = ["Category"] + model_names
    keys    = ["arch_family"] + [metric] * len(model_names)  # align col fmt

    # Format for CSV
    all_rows = data_rows + footer_rows
    saved: dict[str, Path] = {}

    if cfg.export_csv:
        p = _write_csv(all_rows, headers, out / "csv" / f"per_category_{metric}.csv",
                       cfg.missing_csv)
        saved["csv"] = p

    if cfg.export_latex:
        metric_meta = COLUMN_DEFS.get(metric, {})
        cap = caption or (
            f"Per-category {metric_meta.get('header', metric)} "
            f"({metric_meta.get('unit', '')}) across MVTec categories."
        )
        # Build raw latex manually (mixed column types)
        n_models = len(model_names)
        col_spec  = "l" + "r" * n_models
        top    = r"\toprule"    if cfg.booktabs else r"\hline"
        mid    = r"\midrule"    if cfg.booktabs else r"\hline"
        bottom = r"\bottomrule" if cfg.booktabs else r"\hline"
        metric_fmt = metric_meta.get("fmt", cfg.fmt_float)
        better     = metric_meta.get("better")

        # Best per row (best model for each category)
        def _best_col_in_row(row: list[Any]) -> set[int]:
            vals = [(ci, float(v)) for ci, v in enumerate(row[1:], 1)
                    if v is not None and not math.isnan(float(v or "nan"))]
            if not vals:
                return set()
            bv = max(v for _, v in vals) if better == "higher" else min(v for _, v in vals)
            return {ci for ci, v in vals if math.isclose(v, bv, rel_tol=1e-9)}

        body_lines: list[str] = []
        for ri, row in enumerate(data_rows):
            best_cols = _best_col_in_row(row) if cfg.bold_best else set()
            cells = [_latex_escape(str(row[0]))]
            for ci, val in enumerate(row[1:], 1):
                s_val = _fmt_num(val, metric_fmt, cfg.missing)
                if ci in best_cols:
                    s_val = _bold_latex(s_val)
                cells.append(s_val)
            prefix = r"\rowcolor{gray!10}" if cfg.alternating_rows and ri % 2 == 1 else ""
            body_lines.append(prefix + " & ".join(cells) + r" \\")

        footer_lines = [
            " & ".join(
                [_latex_escape(str(fr[0]))]
                + [_fmt_num(v, metric_fmt, cfg.missing) for v in fr[1:]]
            ) + r" \\"
            for fr in footer_rows
        ]

        blocks = [
            rf"\begin{{table}}[{cfg.table_pos}]",
            r"\centering",
        ]
        if cfg.font_size_cmd:
            blocks.append(cfg.font_size_cmd)
        blocks += [
            rf"\caption{{{_latex_escape(cap)}}}",
            rf"\label{{tab:{re.sub(r'[^a-zA-Z0-9:_-]','_',label)}}}"
            if cfg.add_label else "",
            rf"\begin{{tabular}}{{{col_spec}}}",
            top,
            " & ".join(["Category"] + [_latex_escape(mn) for mn in model_names]) + r" \\",
            mid,
        ] + body_lines + [mid] + footer_lines + [
            bottom,
            r"\end{tabular}",
            r"\end{table}",
        ]
        tex = "\n".join(b for b in blocks if b)
        p = out / "latex" / f"per_category_{metric}.tex"
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(tex, encoding="utf-8")
        saved["latex"] = p

    saved["_rows"] = all_rows      # type: ignore[assignment]
    saved["_headers"] = headers    # type: ignore[assignment]
    return saved


def export_statistical_table(
        comparison_report: Any,   # ModelComparisonReport from statistics.py
        *,
        output_dir: Path | str = Path("."),
        caption: str = "Pairwise Wilcoxon signed-rank test results (Bonferroni-corrected).",
        label: str = "statistical_comparison",
        cfg: TableConfig | None = None,
) -> dict[str, Path]:
    """Export pairwise Wilcoxon results and Friedman summary.

    Parameters
    ----------
    comparison_report:
        A ``ModelComparisonReport`` returned by
        :func:`src.evaluation.statistics.compare_models`.
    """
    cfg = cfg or TableConfig()
    out = Path(output_dir)

    rows: list[list[Any]] = []
    headers = ["Model A", "Model B", "W", r"$p$-value", r"$r$", "Sig.", "Interpretation"]
    keys    = ["a", "b", "statistic", "p_value", "effect_size", "significant", "interp"]

    for (a, b), result in comparison_report.pairwise.items():
        sig_mark = r"\checkmark" if result.significant else ""
        rows.append([
            a, b,
            result.statistic,
            result.p_value,
            result.effect_size,
            sig_mark,
            _latex_escape(result.interpretation),
        ])

    saved: dict[str, Path] = {}

    if cfg.export_csv:
        csv_headers = ["Model A", "Model B", "W", "p_value", "r", "Significant", "Interpretation"]
        csv_rows = [
            [r[0], r[1], r[2], r[3], r[4], bool(r[5]), r[6]]
            for r in rows
        ]
        p = _write_csv(csv_rows, csv_headers,
                       out / "csv" / "statistical_comparison.csv", cfg.missing_csv)
        saved["csv"] = p

    if cfg.export_latex:
        # Wilcoxon table
        col_spec = "ll" + "r" * 3 + "c" + "l"
        top    = r"\toprule"    if cfg.booktabs else r"\hline"
        mid    = r"\midrule"    if cfg.booktabs else r"\hline"
        bottom = r"\bottomrule" if cfg.booktabs else r"\hline"

        body: list[str] = []
        for ri, row in enumerate(rows):
            cells = [
                _latex_escape(str(row[0])),
                _latex_escape(str(row[1])),
                _fmt_num(row[2], ".4f", cfg.missing),
                _fmt_num(row[3], ".4f", cfg.missing),
                _fmt_num(row[4], ".3f", cfg.missing),
                str(row[5]),
                str(row[6]),
            ]
            prefix = r"\rowcolor{gray!10}" if cfg.alternating_rows and ri % 2 == 1 else ""
            body.append(prefix + " & ".join(cells) + r" \\")

        # Optional Friedman footer
        fr = comparison_report.friedman
        footer_note = ""
        if fr is not None:
            sig = "significant" if fr.significant else "not significant"
            footer_note = (
                rf"\multicolumn{{{len(headers)}}}{{l}}{{"
                rf"Friedman $\chi^2({fr.df})={fr.statistic:.4f}$, "
                rf"$p={fr.p_value:.4f}$ ({sig})"
                rf".}}"
            )

        blocks = [
            rf"\begin{{table}}[{cfg.table_pos}]",
            r"\centering",
        ]
        if cfg.font_size_cmd:
            blocks.append(cfg.font_size_cmd)
        blocks += [
            rf"\caption{{{_latex_escape(caption)}}}",
            rf"\label{{tab:{re.sub(r'[^a-zA-Z0-9:_-]','_',label)}}}"
            if cfg.add_label else "",
            rf"\begin{{tabular}}{{{col_spec}}}",
            top,
            " & ".join(headers) + r" \\",
            mid,
        ] + body
        if footer_note:
            blocks += [mid, footer_note + r" \\"]
        blocks += [bottom, r"\end{tabular}", r"\end{table}"]
        tex = "\n".join(b for b in blocks if b)
        p = out / "latex" / "statistical_comparison.tex"
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(tex, encoding="utf-8")
        saved["latex"] = p

    saved["_rows"]    = rows     # type: ignore[assignment]
    saved["_headers"] = headers  # type: ignore[assignment]
    return saved


def export_architecture_table(
        candidates: list[dict[str, Any]],
        *,
        col_keys: list[str] | None = None,
        output_dir: Path | str = Path("."),
        caption: str = "Architecture and quantisation configuration of selected candidates.",
        label: str = "architecture_summary",
        cfg: TableConfig | None = None,
) -> dict[str, Path]:
    """Export architecture + quantisation hyperparameter table."""
    cfg = cfg or TableConfig()
    out = Path(output_dir)

    default_keys = [
        "arch_family", "depth", "width",
        "weight_bits", "act_bits", "symmetric", "per_channel",
        "n_params_m", "macs_g",
        "auroc", "latency_ms", "peak_ram_mb", "energy_mj",
    ]
    keys = col_keys or [k for k in default_keys if any(k in c for c in candidates)]
    headers = [_col_header(k) for k in keys]
    rows = [[c.get(k) for k in keys] for c in candidates]

    best_by_col = [_best_row_indices(rows, ci, _col_better(k))
                   for ci, k in enumerate(keys)]

    saved: dict[str, Path] = {}

    if cfg.export_csv:
        p = _write_csv(rows, headers,
                       out / "csv" / "architecture_summary.csv", cfg.missing_csv)
        saved["csv"] = p

    if cfg.export_latex:
        tex = to_latex(candidates, col_keys=keys,
                       caption=caption, label=label, cfg=cfg)
        p = out / "latex" / "architecture_summary.tex"
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(tex, encoding="utf-8")
        saved["latex"] = p

    saved["_rows"]        = rows         # type: ignore[assignment]
    saved["_headers"]     = headers      # type: ignore[assignment]
    saved["_best_by_col"] = best_by_col  # type: ignore[assignment]
    return saved


def export_tracking_table(
        session_results: list[dict[str, Any]],
        *,
        col_keys: list[str] | None = None,
        output_dir: Path | str = Path("."),
        caption: str = "Tracking performance per evaluation scenario.",
        label: str = "tracking_results",
        cfg: TableConfig | None = None,
) -> dict[str, Path]:
    """Export per-scenario tracking performance table.

    Parameters
    ----------
    session_results:
        List of dicts, one per scenario/model pair, with keys such as
        ``scenario``, ``model``, ``iou``, ``precision``, ``recall``,
        ``f1``, ``failure_rate``, ``fps``.
    """
    cfg = cfg or TableConfig()
    out = Path(output_dir)

    default_keys = [
        "scenario", "model",
        "iou", "precision", "recall", "f1",
        "failure_rate", "fps",
    ]
    keys = col_keys or [k for k in default_keys
                        if any(k in r for r in session_results)]
    headers = [_col_header(k) for k in keys]
    rows = [[r.get(k) for k in keys] for r in session_results]

    best_by_col = [_best_row_indices(rows, ci, _col_better(k))
                   for ci, k in enumerate(keys)]

    saved: dict[str, Path] = {}

    if cfg.export_csv:
        p = _write_csv(rows, headers,
                       out / "csv" / "tracking_results.csv", cfg.missing_csv)
        saved["csv"] = p

    if cfg.export_latex:
        tex = to_latex(session_results, col_keys=keys,
                       caption=caption, label=label, cfg=cfg)
        p = out / "latex" / "tracking_results.tex"
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(tex, encoding="utf-8")
        saved["latex"] = p

    saved["_rows"]        = rows         # type: ignore[assignment]
    saved["_headers"]     = headers      # type: ignore[assignment]
    saved["_best_by_col"] = best_by_col  # type: ignore[assignment]
    return saved


def export_repeatability_table(
        repeatability_results: dict[str, Any],
        *,
        output_dir: Path | str = Path("."),
        caption: str = "Repeatability statistics across independent runs.",
        label: str = "repeatability",
        cfg: TableConfig | None = None,
) -> dict[str, Path]:
    """Export repeatability metrics table.

    Parameters
    ----------
    repeatability_results:
        Mapping from metric name → :class:`src.evaluation.statistics.RepeatabilityResult`
        (or equivalent dict with the same fields).
    """
    cfg = cfg or TableConfig()
    out = Path(output_dir)

    keys    = ["metric", "n_runs", "mean", "grand_std", "icc",
               "icc_ci_lo", "icc_ci_hi", "cv", "sem", "mdc95"]
    headers = ["Metric", "Runs", "Mean", "Std",
               "ICC(2,1)", "ICC CI low", "ICC CI high",
               r"CV (\%)", "SEM", r"MDC$_{95}$"]

    rows: list[list[Any]] = []
    for metric_name, res in repeatability_results.items():
        if hasattr(res, "icc"):
            icc_lo, icc_hi = res.icc_ci
            row = [metric_name, res.n_runs, res.mean, res.grand_std,
                   res.icc, icc_lo, icc_hi, res.cv, res.sem, res.mdc95]
        else:
            row = [metric_name] + [res.get(k) for k in keys[1:]]
        rows.append(row)

    best_by_col = [
        _best_row_indices(rows, ci, _col_better(k))
        for ci, k in enumerate(keys)
    ]

    saved: dict[str, Path] = {}

    if cfg.export_csv:
        csv_headers = ["Metric", "n_runs", "mean", "std",
                       "icc", "icc_ci_lo", "icc_ci_hi", "cv_pct", "sem", "mdc95"]
        p = _write_csv(rows, csv_headers,
                       out / "csv" / "repeatability.csv", cfg.missing_csv)
        saved["csv"] = p

    if cfg.export_latex:
        col_spec = "l" + "r" * (len(keys) - 1)
        top    = r"\toprule"    if cfg.booktabs else r"\hline"
        mid    = r"\midrule"    if cfg.booktabs else r"\hline"
        bottom = r"\bottomrule" if cfg.booktabs else r"\hline"

        body: list[str] = []
        for ri, row in enumerate(rows):
            cells = [_latex_escape(str(row[0]))]
            for ci in range(1, len(keys)):
                fmt = COLUMN_DEFS.get(keys[ci], {}).get("fmt", cfg.fmt_float)
                val_str = _fmt_num(row[ci], fmt, cfg.missing)
                if cfg.bold_best and ri in best_by_col[ci]:
                    val_str = _bold_latex(val_str)
                cells.append(val_str)
            prefix = r"\rowcolor{gray!10}" if cfg.alternating_rows and ri % 2 == 1 else ""
            body.append(prefix + " & ".join(cells) + r" \\")

        blocks = [
            rf"\begin{{table}}[{cfg.table_pos}]",
            r"\centering",
        ]
        if cfg.font_size_cmd:
            blocks.append(cfg.font_size_cmd)
        blocks += [
            rf"\caption{{{_latex_escape(caption)}}}",
            rf"\label{{tab:{re.sub(r'[^a-zA-Z0-9:_-]','_',label)}}}"
            if cfg.add_label else "",
            rf"\begin{{tabular}}{{{col_spec}}}",
            top,
            " & ".join(headers) + r" \\",
            mid,
        ] + body + [bottom, r"\end{tabular}", r"\end{table}"]
        tex = "\n".join(b for b in blocks if b)
        p = out / "latex" / "repeatability.tex"
        p.parent.mkdir(parents=True, exist_ok=True)
        p.write_text(tex, encoding="utf-8")
        saved["latex"] = p

    saved["_rows"]    = rows     # type: ignore[assignment]
    saved["_headers"] = headers  # type: ignore[assignment]
    return saved


# ---------------------------------------------------------------------------
# Excel workbook writer (all tables in one file)
# ---------------------------------------------------------------------------

def _write_all_excel(
        all_tables: dict[str, dict[str, Any]],
        path: Path,
        cfg: TableConfig,
) -> Path:
    """Write every table into a single Excel workbook, one sheet each."""
    if not _HAVE_OPENPYXL:
        warnings.warn(
            "openpyxl not installed; Excel export skipped. "
            "Install it with: pip install openpyxl",
            ImportWarning, stacklevel=2,
        )
        return path

    import openpyxl as _xl
    wb = _xl.Workbook()
    wb.remove(wb.active)  # remove default empty sheet

    for sheet_name, table in all_tables.items():
        rows    = table.get("_rows", [])
        headers = table.get("_headers", [])
        keys    = table.get("_keys", [])
        best_by_col = table.get("_best_by_col")
        if not rows:
            continue
        _write_excel_sheet(wb, sheet_name, rows, headers, keys or None,
                           cfg, best_by_col)

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    LOG.info("Excel workbook → %s", path)
    return path


# ---------------------------------------------------------------------------
# Orchestrator
# ---------------------------------------------------------------------------

def export_all_tables(
        output_dir: Path | str,
        *,
        candidates: list[dict[str, Any]] | None = None,
        results_by_model: dict[str, dict[str, dict[str, Any]]] | None = None,
        comparison_report: Any | None = None,
        session_results: list[dict[str, Any]] | None = None,
        repeatability_results: dict[str, Any] | None = None,
        categories: list[str] | None = None,
        metrics: list[str] | None = None,
        cfg: TableConfig | None = None,
) -> dict[str, Path]:
    """Generate all standard thesis tables and write them to *output_dir*.

    Parameters
    ----------
    output_dir:
        Root directory; ``csv/``, ``latex/``, and ``excel/`` sub-directories
        are created automatically.
    candidates:
        Full NSGA-II population (must include ``is_pareto``).
    results_by_model:
        ``{model_name: {category: metrics_dict}}`` from test_metrics.py.
    comparison_report:
        :class:`src.evaluation.statistics.ModelComparisonReport`.
    session_results:
        List of tracking session dicts from main_tracking.py.
    repeatability_results:
        ``{metric_name: RepeatabilityResult}`` from statistics.py.
    categories:
        Ordered list of MVTec categories for the per-category table.
    metrics:
        Which per-category metrics to export (default: all four objectives).
    cfg:
        Formatting configuration (shared across all tables).

    Returns
    -------
    dict[str, Path]
        ``{table_name: Path}`` for every file written.
    """
    cfg     = cfg or TableConfig()
    out     = Path(output_dir)
    metrics = metrics or ["auroc", "pixel_auroc", "pr_auc"]
    all_tables: dict[str, dict[str, Any]] = {}
    saved:      dict[str, Path] = {}

    def _merge(name: str, result: dict[str, Any]) -> None:
        all_tables[name] = result
        for fmt_key in ("csv", "latex"):
            if fmt_key in result and isinstance(result[fmt_key], Path):
                saved[f"{name}_{fmt_key}"] = result[fmt_key]

    if candidates:
        _merge("pareto",       export_pareto_table(candidates, output_dir=out, cfg=cfg))
        _merge("architecture", export_architecture_table(candidates, output_dir=out, cfg=cfg))

    if results_by_model:
        for metric in metrics:
            if any(
                any(metric in cat_res for cat_res in model_res.values())
                for model_res in results_by_model.values()
            ):
                key = f"per_category_{metric}"
                _merge(key, export_per_category_table(
                    results_by_model, categories=categories,
                    metric=metric, output_dir=out, cfg=cfg,
                ))

    if comparison_report is not None:
        _merge("statistical", export_statistical_table(
            comparison_report, output_dir=out, cfg=cfg,
        ))

    if session_results:
        _merge("tracking", export_tracking_table(
            session_results, output_dir=out, cfg=cfg,
        ))

    if repeatability_results:
        _merge("repeatability", export_repeatability_table(
            repeatability_results, output_dir=out, cfg=cfg,
        ))

    # Combined Excel workbook
    if cfg.export_excel:
        xl_path = _write_all_excel(
            all_tables, out / "excel" / "all_tables.xlsx", cfg,
        )
        saved["excel_workbook"] = xl_path

    LOG.info("export_all_tables: wrote %d files to %s", len(saved), out)
    return saved
